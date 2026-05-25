import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def generar_excel(fecha, cajero, total, efectivo, tarjeta, transferencia, archivo_salida):
    try:
        carpeta = os.path.dirname(archivo_salida)
        if carpeta and not os.path.exists(carpeta):
            os.makedirs(carpeta, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cierre"

        encabezados = ["Fecha", "Cajero", "Total", "Efectivo", "Tarjeta", "Transferencia", "FechaHora"]
        datos = [fecha, cajero, total, efectivo, tarjeta, transferencia, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]

        fill_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font_encabezado = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        font_datos = Font(name="Calibri", size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=1, column=col_idx, value=encabezado)
            celda.fill = fill_encabezado
            celda.font = font_encabezado
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = border

        total_row = 2
        if isinstance(efectivo, (int, float)):
            total_row = efectivo + tarjeta + transferencia + 2

        for col_idx, valor in enumerate(datos, start=1):
            celda = ws.cell(row=2, column=col_idx, value=valor)
            celda.font = font_datos
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = border

        for col_idx in range(1, len(encabezados) + 1):
            max_len = len(str(encabezados[col_idx - 1]))
            valor = datos[col_idx - 1]
            if valor is not None:
                max_len = max(max_len, len(str(valor)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        wb.save(archivo_salida)
        logger.info("Excel generado exitosamente: %s", archivo_salida)
        return archivo_salida

    except Exception:
        logger.exception("Error al generar el archivo Excel")
        raise
